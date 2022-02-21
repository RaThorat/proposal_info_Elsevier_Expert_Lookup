# -*- coding: utf-8 -*-
"""
Created on Fri Jan 14 12:56:07 2022

@author: r.thorat
"""
#importing necessary modules
import pandas._libs.tslibs.base
import os
os.getcwd()
from tkinter import filedialog as fd
import pandas as pd
import re
from nameparser import HumanName
import os
os.getcwd()

#define functions
#name of the round
def sub_round():

    """ callback when the button is clicked
    """
    msg = f'You entered: {round.get()}'
    showinfo(
        title='Information',
        message=msg
    )
    return round.get()
#to get suggested reviewers, if any    
def getsuggestedreviewers0():
    global df_SR_0
    import_file_path2 = fd.askopenfilename()
    df_SR_0 = pd.read_excel (import_file_path2)
    return df_SR_0

#to get country sheet
def country_sheet():
    global df_CS
    import_file_path7 = fd.askopenfilename()
    df_CS = pd.read_excel(import_file_path7)
    return df_CS

# Create Tkinter GUI to interact with the code
##https://www.geeksforgeeks.org/tkinter-application-to-switch-between-different-page-frames/
import tkinter as tk
from tkinter import ttk
from tkinter.messagebox import showinfo

def Close():
    app.destroy() 
         
LARGEFONT =("Verdana", 35)

class tkinterApp(tk.Tk):
	
	# __init__ function for class tkinterApp
	def __init__(self, *args, **kwargs):
		
		# __init__ function for class Tk
		tk.Tk.__init__(self, *args, **kwargs)
		
		# creating a container
		container = tk.Frame(self)
		container.pack(side = "top", fill = "both", expand = True)

		container.grid_rowconfigure(0, weight = 1)
		container.grid_columnconfigure(0, weight = 1)

		# initializing frames to an empty array
		self.frames = {}
        

        
		# iterating through a tuple consisting
		# of the different page layouts
		for F in (StartPage, Page1):

			frame = F(container, self)

			# initializing frame of that object from
			# startpage, page1, page2 respectively with
			# for loop
			self.frames[F] = frame

			frame.grid(row = 0, column = 0, sticky ="nsew")
            


		self.show_frame(StartPage)

	# to display the current frame passed as
	# parameter
	def show_frame(self, cont):
		frame = self.frames[cont]
		frame.tkraise()
   
     
# first window frame startpage

class StartPage(tk.Frame):
	def __init__(self, parent, controller):
		tk.Frame.__init__(self, parent)
		
		# label of frame Layout 2
		label = ttk.Label(self, text ="Startpage")
		
		# putting the grid in its place by using
		# grid
		label.grid(row = 0, column = 1, padx = 10, pady = 10)

		button1 = ttk.Button(self, text ="Info for Expert Lookup", 
		command = lambda : controller.show_frame(Page1))
	
		# putting the button in its place by
		# using grid
		button1.grid(row = 1, column = 1,  padx = 10, pady = 10)
        # button to show proposalinfo with text
		# layout3
		button2 = ttk.Button(self, text ="Close",
							command =Close)
	
		# putting the button in its place by
		# using grid
		button2.grid(row = 2, column = 1, padx = 10, pady = 10)

# fourth window frame page1
class Page1(tk.Frame):
	def __init__(self, parent, controller):
		tk.Frame.__init__(self, parent)
		label = ttk.Label(self, text ="Info for Expert Lookup",  font = LARGEFONT)
		label.grid(row = 0, column = 1, padx = 40, pady = 10)

		# button to show frame 3 with text
		# layout3
		button3 = ttk.Button(self, text ="Back to Startpage",
							command = lambda : controller.show_frame(StartPage))
	
		# putting the button in its place by
		# using grid
		button3.grid(row = 1, column = 1, padx = 10, pady = 10)
        
        
        

		my_label = ttk.Label(self, text ="Enter subsidy name (example: \'%XXX-Talentprogramma XXX XXX 2021%\'):")
        
	
		# putting the button in its place by
		# using grid
		my_label.grid(row = 1, column = 2, padx =60, pady = 10) 
        # putting the button in its place by
		# using grid
		global round 
        
        
		round = tk.StringVar(self) 
        
        # button to show proposalinfo with text
		# layout3
		my_box = ttk.Entry(self, textvariable=round, width=60)
        # putting the button in its place by
		# using grid
		my_box.grid(row = 2, column = 2, padx = 10, pady = 10)
        
        # button to show proposalinfo with text
		# layout3
		button4 = ttk.Button(self, text ="Click me to confirm",
							command =sub_round)
	
		# putting the button in its place by
		# using grid
		button4.grid(row = 3, column = 2, padx = 10, pady = 10)
        
        
        # button to show suggested reviewers with text
		# layout3
		button5 = ttk.Button(self, text ="Import sugg reviewers",
							command =getsuggestedreviewers0)
	
		# putting the button in its place by
		# using grid
		button5.grid(row = 4, column = 2, padx = 10, pady = 10)
        
        
        
        # button to input export file name with text
		# layout3
		button9 = ttk.Button(self, text ="Import country sheet",
							command =country_sheet)
	
		# putting the button in its place by
		# using grid
		button9.grid(row = 6, column = 2, padx = 10, pady = 10)
        # button to show proposalinfo with text
		# layout3
		button8 = ttk.Button(self, text ="Close",
							command =Close)
	
		# putting the button in its place by
		# using grid
		button8.grid(row = 7, column = 1, padx = 10, pady = 10)
# Driver Code
app = tkinterApp()
#https://newbedev.com/how-to-change-the-color-of-ttk-button
style = ttk.Style()
style.theme_use('alt')
style.configure('TButton', background = '#232323', foreground = 'white', width = 20, borderwidth=1, focusthickness=3, focuscolor='none')
style.map('TButton', background=[('active','#ff0000')])
app.title("This is my Interface")
#app.bind('<Return>', run_and_close)
#app.bind('<Escape>', close)

app.mainloop()

#Write query in SQL service managemnt studio to get required information such as
#File number,Main applicant,Main applicant_Last name,First name,Gender,Promotion date,C_Email,
#Main Organization,C_Organization,C_Postal code,C_City,Title,Summary,Main discipline,Subdiscipline,Word

query_code='YOUR CODE where sub.subsidieronde_naam like'

Sub_Naam=round.get()
Cond='and rol.Naam in' + ' (' + "'Hoofdaanvrager','Medeaanvrager'"+')'
query=query_code+' '+Sub_Naam + Cond
#importing necessary modules
print(query)
import pyodbc 

conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=YOUR SERVER;'
                      'Database=YOUR DATABASE;'
                      'Trusted_Connection=yes;')
#creation of a dataframe
df_pv = pd.read_sql_query(query, conn) 
df_pv1=df_pv[df_pv['rolnaam']=='Hoofdaanvrager']
df_pv2 = df_pv.groupby(
        ['dossiernummer', 'Hoofdaanvrager', 'achternaam',
           'voornaam', 'Geslacht', 'Promotiedatum', 'C_Email',
           'Correspondentietaal', 'Hoofdorganisatie', 'C_Organisatie',
           'C_Postcode', 'C_Plaats', 'Titel', 'Samenvatting', 'Hoofddiscipline',
           'Subdiscipline','rolnaam', 'subsidieronde_naam', 'atl_Ingetrokken',
           'atl_Gehonoreerd'],
    )['woord'].agg(', '.join).reset_index(name='Trefwoorden')
df_pv2.shape
#If you want the proposal info for some other reason, you can export it as excel file



# # Stap 2 Edit information in terms of 'Proposal Upload Example.xlsx' from expert look up
#2.a Formulating information for 'Proposal Details' worksheet
df_PD= pd.DataFrame(columns = ['No', 'Grant No.', 'Council', 'Proposal Title', 'Abstract', 'Specific Aims', 'Keywords'])

#Counting the proposals and creating number column
#df_PD['Proposal Number']=np.arange(len(df_pv2)) 

#Changing names of the columns
df_PD['Grant No.']=df_pv2['dossiernummer']
df_PD['No'] = df_pv2['dossiernummer'].str[-3:]
df_PD['No'] = df_PD['No'].astype(int)
df_PD['Council']='Name of the funder'
df_PD['Proposal Title']=df_pv2['Titel']
df_PD['Abstract']=df_pv2['Samenvatting']
df_PD['Keywords']=df_pv2['Trefwoorden']
#check
df_PD.head()
# 2.b Formulate information for 'Proposal Applicants' worksheet, main applicant

df_PA= pd.DataFrame(columns = ['Proposal Number', 'Grant No.', 'Last Name', 'First Name',
                               'Scopus Author ID', 'OrcId', 'Email', 'Affiliation', 'Country', 'Role', 'Is Principal Investigator?'])
#Changing names of the columns
df_PA['Grant No.']=df_pv2['dossiernummer']
df_PA['Proposal Number'] = df_pv2['dossiernummer'].str[-3:]
df_PA['Proposal Number'] = df_PA['Proposal Number'].astype(int)
df_PA['Last Name']=df_pv2['achternaam']
df_PA['First Name']=df_pv2['voornaam']
df_PA['Email']=df_pv2['C_Email']
df_PA['Affiliation']=df_pv2['Hoofdorganisatie']
df_PA['Country']='Netherlands'
df_PA['Role']='Principal Investigator'
df_PA['Is Principal Investigator?']='Yes'

#check
df_PA.head(2)

# 2.b.ii Formulate information for 'Proposal Applicants' worksheet, co-applicant

df_colab_0=df_pv[df_pv['rolnaam']=='Medeaanvrager']
df_colab_1 = df_colab_0.groupby(
        ['dossiernummer', 'achternaam',
           'voornaam', 'Geslacht', 'Promotiedatum', 'C_Email',
           'Correspondentietaal', 'Hoofdorganisatie', 'C_Organisatie',
           'C_Postcode', 'C_Plaats', 'Titel', 'Samenvatting', 'Hoofddiscipline',
           'Subdiscipline','rolnaam', 'subsidieronde_naam', 'atl_Ingetrokken',
           'atl_Gehonoreerd'],
    )['woord'].agg(', '.join).reset_index(name='Trefwoorden')
df_pv2.shape
df_colab= pd.DataFrame(columns = ['Proposal Number', 'Grant No.', 'Last Name', 'First Name',
                               'Scopus Author ID', 'OrcId', 'Email', 'Affiliation', 'Country', 'Role', 'Is Principal Investigator?'])
#Changing names of the columns

df_colab['Grant No.']=df_colab_1['dossiernummer']
df_colab['Proposal Number'] = df_colab_1['dossiernummer'].str[-3:]
df_colab['Proposal Number'] = df_colab['Proposal Number'].astype(int)
df_colab['Last Name']=df_colab_1['achternaam']
df_colab['First Name']=df_colab_1['voornaam']
df_colab['Email']=df_colab_1['C_Email']
df_colab['Affiliation']=df_colab_1['Hoofdorganisatie']
df_colab['Country']=''
df_colab['Role']='Co-Investigator'
df_colab['Is Principal Investigator?']='No'


#adding frame
frames = [df_PA, df_colab]


#concating the dataframe
df_PA3 = pd.concat(frames, sort=False)

#check
df_PA3.head(2)

df_PA33=df_PA3.reset_index(drop=True)


# 2.c Call up the excel sheets 'Suggested reviewers' and 'country sheet'


df_SR=pd.DataFrame(columns=['Proposal Number','Grant No.','Last Name', 'First Name', 'Scopus Author ID', 'OrcId', 'Email', 'Affiliation', 'Country'])

#seperating the reviewer's name from the text
if len(df_SR_0)!=0:
    df_SR_1=df_SR_0
    df_SR_1['referent naam']=df_SR_1['Voorgesteld referent'].str.split(',').str[0]
    df_SR_1.head()
    First_Name=[]
    Last_Name=[]
    for person in df_SR_1['referent naam']:
        name = HumanName(person)
        First_Name.append(name.first)
        Last_Name.append(name.last)
    df_SR_1['First Name'] = First_Name
    df_SR_1['Last Name'] = Last_Name
    #pd.DataFrame({'First_Name':First_Name,'Last_Name':Last_Name})
    #Extraction of email adress
    email=[]
    for referenten in df_SR_1['Voorgesteld referent']:
        # \w matches any non-whitespace character
        # @ for as in the Email
        # + for Repeats a character one or more times
        FindEmail = re.findall("([\w.-]+@[\w.-]+)", referenten)
        # run for loop on the list variable
        for l in FindEmail:
        #set value in email variable
            emaill=l
        #declare local variables to store email addresses
        email.append(emaill)
    df_SR_1['Email']=email
    #extraction of university name
    df_uni = pd.DataFrame()
    for referenten in df_SR_1['Voorgesteld referent']:
        text=referenten
        matchlist = ['Hospital','University','Universitäts','Universität','Università','Hogeschool','Labs', 'Laboratory', 'Zoo','Institute','Institut','School','Ecole','Academy', 'Universite','College','Universitaet,' '* University'] 
        # define all keywords that you need look up
        p = re.compile('^(.*?),\s+(.*?),(.*?)\.')   # regex pattern to match
        # We use a list comprehension using 'any' function to check if any of the item in the matchlist can be found in either group1 or group2 of the pattern match results
        text_match = [m.group(1) if any(x in m.group(1) for x in matchlist) else m.group(2) for m in re.finditer(p,text)]
        df_uni = df_uni.append(text_match, ignore_index=True)
    df_SR_1['Affiliation']=df_uni.iloc[0:]
    #Extracting last digits from the grant number, to fill in as proposal number. 
    df_SR['Proposal Number']=df_SR_1['Project dossier'].str.split('.').str[-1]
    df_SR['Proposal Number'] = df_SR['Proposal Number'].astype(int)
    df_SR['Grant No.']=df_SR_1['Project dossier']
    df_SR['Last Name']=df_SR_1['Last Name']
    df_SR['First Name']=df_SR_1['First Name']
    df_SR['Scopus Author ID']='NaN'
    df_SR['OrcId']='NaN'
    df_SR['Email']=df_SR_1['Email']
    df_SR['Affiliation']=df_SR_1['Affiliation']
    df_SR['Country']='NaN'

# # Stap 3 Bundling requested information in one excel workbook and download it



#The openpyxl.utils.dataframe.dataframe_to_rows() function provides a simple way to work with Pandas Dataframes:

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook


wb = Workbook()
ws = wb.active
 
ws.title='Proposal Details'

for r in dataframe_to_rows(df_PD, index=False, header=True):
    ws.append(r)

for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'

#writing proposal applicants
ws1 = wb.create_sheet("Proposal Applicants", 1) 
for r in dataframe_to_rows(df_PA33, index=False, header=True):
    ws1.append(r)

#writing suggested reviewers
ws2 = wb.create_sheet("Suggested Reviewers", 2) 
for r in dataframe_to_rows(df_SR, index=False, header=True):
    ws2.append(r)

#writing country sheets
ws3 = wb.create_sheet("country sheet", 3) 
for r in dataframe_to_rows(df_CS, index=False, header=True):
    ws3.append(r)
    
#saving the excel workbook
Exportfile_name=('Proposal info to feed in EL.xlsx ')
wb.save(Exportfile_name)